import frappe
from frappe.utils import formatdate, format_datetime
from frappe.utils.file_manager import get_file

import csv
import io
from openpyxl import load_workbook


@frappe.whitelist()
def get_scholar_field_property(scholar, fieldname):
    if not (scholar and fieldname):
        return

    field = frappe.get_meta("Scholar").get_field(fieldname)
    if not field:
        return

    value = frappe.db.get_value("Scholar", scholar, fieldname)

    if field.fieldtype == "Date":
        value = formatdate(value)
    elif field.fieldtype == "Datetime":
        value = format_datetime(value)

    return {
        "value": value,
        "datatype": field.fieldtype,
        "label": field.label,
        "options": field.options,
    }


def extract_data_from_file(file_url):
    file_doc = get_file(file_url)
    filename, content = file_doc[0], file_doc[1]
    rows = []

    if filename.lower().endswith(".csv"):
        try:
            stream = io.StringIO(content)
            reader = csv.DictReader(stream)
            rows = list(reader)
        except Exception as e:
            frappe.throw(f"Failed to read CSV: {str(e)}")

    elif filename.lower().endswith((".xlsx", ".xls")):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [
                str(cell.value).strip() for cell in ws[1] if cell.value is not None
            ]

            for row_data in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row_data):
                    rows.append(dict(zip(headers, row_data)))
        except Exception as e:
            frappe.throw(f"Failed to read Excel: {str(e)}")
    else:
        frappe.throw("Unsupported file type. Upload CSV or Excel.")

    return rows


def get_doctype_headers(doctype: str) -> list[str]:
    SYSTEM_FIELDS = {
        "name",
        "parent",
        "parentfield",
        "parenttype",
        "idx",
        "creation",
        "modified",
        "owner",
        "docstatus",
    }
    meta = frappe.get_meta(doctype)
    return [
        f.fieldname
        for f in meta.fields
        if f.fieldtype not in ("Section Break", "Column Break", "HTML", "Table")
        and f.fieldname not in SYSTEM_FIELDS
    ]
